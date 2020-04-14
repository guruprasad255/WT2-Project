from flask import Flask, render_template,jsonify,request,abort,Response,json,_request_ctx_stack,redirect,url_for
import requests
from flask_cors import cross_origin
import sqlite3
import re
import datetime
import csv
import random 
from flask_cors import CORS
import openpyxl
from plotly.offline import plot
from plotly.graph_objs import Bar,Pie
from flask import Markup
from flask import Flask, render_template,request,session
import pandas as pd
import os
import plotly.graph_objects as go
import plotly.io  as pio
#app=Flask(__name__)
#l=os.getcwd()
app = Flask(__name__,template_folder='/var/www/html/new/')
cors = CORS(app, resources={r"/api/*": {"origins": "*"}})
cors = CORS(app, resources={r"/rcreg/*": {"origins": "*"}})
cors = CORS(app, resources={r"/signup": {"origins": "*"}})
cors = CORS(app, resources={r"/ulogin": {"origins": "*"}})
cors = CORS(app, resources={r"/ulogout": {"origins": "*"}})
cors = CORS(app, resources={r"/alogin": {"origins": "*"}})
cors = CORS(app, resources={r"/alogout": {"origins": "*"}})
cors = CORS(app, resources={r"localhost/new/*":{"origins":"*"}})
cors = CORS(app, resources={r"/approval/*": {"origins": "*"}})
flag = 0
ll_applications = 0
dl_applications = 0
rc_applications = 0
srev = 0 #dl ll
vrev = 0 #rc
  
@app.route("/api/v1/Revenue",methods=["GET"])
@cross_origin(origin="*")
def Revenue():
    if(request.method == 'GET'):
        
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":3})
        if(readresponse.status_code == 204):
            return jsoinfy({}),400
        #print("Count here",readresponse.json())
        response = app.response_class(response=json.dumps(readresponse.json()),status=200,mimetype='application/json')
        return response
    else:
        return Response(status=405)

@app.route("/api/v1/RCReg",methods=["POST","OPTIONS"])
@cross_origin(origin='*')
def insertIntoRCReg():
    reg_id = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"Count","column":"reg_id","src":"","dst":"","where":0})
    reg_id = reg_id.json()[0][0]
    
    if(request.method == 'POST'):
        try:
            reg = str(reg_id)
            RTO = request.get_json()["rto"]
            chassis = request.get_json()["chassis_num"]
            Engine = request.get_json()["engine_num"]
            Owner = request.get_json()["name"]
            Address = request.get_json()["address"]
            Mobile = request.get_json()["phno"]
            Manufacturer = request.get_json()["manufacturer"]
            Model = request.get_json()["model"]
            Color = request.get_json()["color"]
            Class = request.get_json()["class"]
            DOR = str(datetime.datetime.now())
            Approved = "0"
            Vehicle_no = RTO+" "+str(random.randint(1000,9999))
            
            
        except KeyError:
            print("Am I here ?	")
            return jsonify({}),400
        
        detlist = [reg,RTO,chassis,Engine,Owner,Address,Mobile,Manufacturer,Model,Color,Class,DOR,Approved,Vehicle_no]
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"RCReg","column":"Engine","src":"Engine","dst":Engine,"where":1})
        
        if(readresponse.status_code==204):
            writeresponse = requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"RCReg","value":detlist})
            if(writeresponse.status_code == 201):
                global rc_applications
                rc_applications = 1
                global vrev 
                vrev = 1
                global ll_applications
                ll_applications = 0
                global dl_applictions
                dl_applications = 0
                 
                list_revenue = [RTO,DOR,str(srev),str(vrev)]
                list_total = [DOR,RTO,str(ll_applications),str(dl_applications),str(rc_applications)]
                requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"Revenue","value":list_revenue})
                requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"Total","column":"RTO","value":list_total})
                requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":2,"column":"reg_id"})
                xfile = openpyxl.load_workbook('regdata.xlsx')
                sheet = xfile.get_sheet_by_name('regdata.xlsx')
                value = sheet['C53'].value
                sheet['C53']=value+1
                xfile.save('regdata.xlsx')
                xfile = openpyxl.load_workbook('vahanrevenue.xlsx')
                sheet = xfile.get_sheet_by_name('Sheet1')
                value = sheet['C53'].value
                sheet['C53']=value+1
                xfile.save('vahanrevenue.xlsx')
                return jsonify({}),201
            elif(writeresponse.status_code == 500):
                return jsonify({}),500
        else:
            return jsonify({}),400
        
    else:
        return Response(status=405)
    

@app.route("/api/v1/LLReg",methods=["POST"])
@cross_origin(origin="*")
def insertIntoLLReg():
    
    
    ll_reg = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"Count","column":"ll_reg","src":"","dst":"","where":0})
    ll_reg =  ll_reg.json()[0][0]
    						
    if(request.method == 'POST'):
        try:
            reg_id = str(ll_reg)
            RTO = request.get_json()["rto"]
            Pincode = request.get_json()["pin"]
            Name = request.get_json()["name"]
            DOB = request.get_json()["DoB"]
            Gender = request.get_json()["gender"]
            Age = str(request.get_json()["age"])
            Mail = request.get_json()["email"]
            Mobile = request.get_json()["phno"]
            Address = request.get_json()["address"]
            Class = str(request.get_json()["cls"])
            Approved = "0"
            DOA = str(datetime.datetime.now())

            print(DOB)     
        except KeyError:
            
            return jsonify({}),400
        
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":2,"table":"LLReg","column":"Name,Mail,DOB","src":["Name","Mail","DOB"],"dst":[Name,Mail,DOB],"where":1})
        
        if(readresponse.status_code == 204):
                 
        
            detlist = [reg_id,RTO,Pincode,Name,DOB,Gender,Age,Mail,Mobile,Address,Class,Approved,DOA]
            writeresponse = requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"LLReg","value":detlist})
            if(writeresponse.status_code == 201):
                global ll_applications
                ll_applications = 1
                global srev
                srev = 1
                global rc_applications
                rc_applications = 0
                global dl_applictions
                dl_applications = 0
                list_revenue = [RTO,DOA,str(srev),str(vrev)]
                list_total = [DOA,RTO,str(ll_applications),str(dl_applications),str(rc_applications)]
                requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"Revenue","value":list_revenue})
                requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"Total","value":list_total})
                requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":2,"column":"ll_reg"})
                xfile = openpyxl.load_workbook('applcount.xlsx')
                sheet = xfile.get_sheet_by_name('Sheet1')
                value = sheet['B2'].value
                sheet['B2']=value+1
                xfile.save('applcount.xlsx')
                return jsonify(ll_reg),201
            elif(writeresponse.status_code == 500):
                return jsonify({}),500
        else:	
            
            return jsonify({}),400
            


    else:
        return Response(status=405)


@app.route("/api/v1/DLReg",methods=["POST","OPTIONS"])
@cross_origin(origin="*")
def insertIntoDLReg():
    
    if(request.method == 'POST'):
        try:
            LLno = request.get_json()["llnum"]
            Dob = request.get_json()["DoB"]
        except KeyError:
            return jsonify({}),400
        
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"DLReg","column":"LLno","src":"LLno","dst":LLno,"where":1})
        if(readresponse.status_code != 204):
            return "already applied"


        RTO = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"LLReg","column":"RTO","src":"reg_id","dst":LLno,"where":1})
        if(RTO.status_code ==204):
            return jsonify({}),400
        RTO =RTO.json()
           
        readresponse_llnum = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"LLReg","column":"reg_id","src":"reg_id","dst":LLno,"where":1})
            
        if(readresponse_llnum.status_code == 204):
            return "invalid ll number"
            
        readresponse_dob = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"LLReg","column":"DOB","src":"reg_id","dst":LLno,"where":1})
        if(readresponse_dob.status_code==204):
            return "invalid dob"
        for i in readresponse_dob.json():
            if(i[0]==Dob):
                #Dob is matching
                DLno =str(RTO[0][0])+" "+str(random.randint(10000000000,99999999999))
                Validity = str(datetime.datetime.now() + datetime.timedelta(days=365*20))
                DOA = str(datetime.datetime.now())
                Flag = "0" 
                Status = "0"
            else:
                return "invalid dob"
        
        
        detlist = [LLno,DLno,Validity,DOA,Flag,Status]
        #print(detlist)
        writeresponse = requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"DLReg","value":detlist})
        if(writeresponse.status_code == 201):
            print("Am I here twice ?")
            global srev
            srev = 1
            global dl_applications
            dl_applications = 1
            global ll_applications
            ll_applications = 0
            global rc_applictions
            rc_applicatiions = 0
            list_revenue = [str(RTO[0][0]),DOA,str(srev),str(vrev)]
            list_total = [DOA,str(RTO[0][0]),str(ll_applications),str(dl_applications),str(rc_applications)] 
            requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"Revenue","value":list_revenue})
            requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"Total","value":list_total})
            xfile = openpyxl.load_workbook('applcount.xlsx')
            sheet = xfile.get_sheet_by_name('Sheet1')
            value = sheet['B3'].value
            sheet['B3']=value+1
            xfile.save('applcount.xlsx')
            return DLno
        elif(writeresponse.status_code == 500):
            return jsonify({}),500
    else:
        return Response(status=405)


@app.route("/api/v1/Total",methods=["GET"])
def Total():
    if(request.method == 'GET'):
        
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":4})
        if(readresponse.status_code == 204):
            return jsoinfy({}),400
        response = app.response_class(response=json.dumps(readresponse.json()),status=200,mimetype='application/json')
        return response
    else:
        return Response(status=405)


@app.route('/signup', methods=['GET', 'POST'])
@cross_origin(origin="*")
def signup():
    if request.method == 'POST':
            uname = request.get_json()["uname"]
            password = request.get_json()["password"]
            rpassword = request.get_json()["rpassword"]
            if(password==rpassword):
                detlist = [uname,password]
                readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"Users","column":"username","src":"username","dst":uname,"where":1})
                if(readresponse.status_code==204):
                    writeresponse = requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":1,"table":"Users","value":detlist})
                    if(writeresponse.status_code == 201):
                        return "created"
                    elif(writeresponse.status_code == 500):
                        return jsonify({}),500
                else:
                    return "user already exists"
            else:
                return "passwords don't match"

@app.route('/ulogin', methods=['GET', 'POST'])
@cross_origin(origin="*")
def ulogin():
    if request.method == 'POST':
            uname = request.get_json()["uname"]
            password = request.get_json()["password"]
            readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"Users","column":"*","src":"username","dst":uname,"where":1})
            if(readresponse.status_code==204):
                return Response(status=210)
            else:
                record=readresponse.json()
                if(record[0][1]==password):
                    session["username"]=uname
                    return Response(status=200)
                else:
                    return Response(status=211)


@app.route('/alogin', methods=['GET', 'POST'])
@cross_origin(origin="*")
def alogin():
    if request.method == 'POST':
            uname = request.get_json()["uname"]
            password = request.get_json()["password"]
            readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"Admin","column":"*","src":"username","dst":uname,"where":1})
            if(readresponse.status_code==204):
                return Response(status=210)
            else:
                record=readresponse.json()
                if(record[0][1]==password):
                    print(record[0][1])
                    session["adminuser"]=uname
                    return Response(status=200)
                else:
                    return Response(status=211)
            

@app.route('/approval/dl/<num>', methods=['GET', 'POST'])
@cross_origin(origin="*")
def dlapproval(num):
    if request.method == 'GET':
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"DLReg","column":"DLno","src":"DLno","dst":num,"where":1})
        if(readresponse.status_code == 204):
            return "Invalid DL"
        requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":3,"value":num})
        return "approved"
        
            
@app.route('/approval/ll/<num>', methods=['GET', 'POST'])           
@cross_origin(origin="*")
def llapproval(num):
    if request.method == 'GET':
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"LLReg","column":"reg_id","src":"reg_id","dst":num,"where":1})
        if(readresponse.status_code == 204):
            return "Invalid LL"
        requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":5,"value":num})
        return "approved"
 
@app.route('/approval/rc/<num>', methods=['GET', 'POST'])
@cross_origin(origin="*")
def rcapproval(num):
    if request.method == 'GET':
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"RCReg","column":"vahicle_no","src":"vahicle_no","dst":num,"where":1})
        if(readresponse.status_code == 204):
            return "Invalid RC"
        requests.post("http://127.0.0.1:5000/api/v1/db/write",json={"OP":4,"value":num})
        return "approved"
 





@app.route("/api/v1/db/write",methods=["POST"])
@cross_origin(origin="*")
def writeDB():
    OP = request.get_json()["OP"]
    try:
        conn = sqlite3.connect("webdatabase.db")
     
         
    except:
        return Response(status=500)

    if(OP == 1):# INSERT VALUES INTO ANY TABLE 
        table = request.get_json()['table']
        value = request.get_json()['value']
        final_values=''
        for i in value:
            final_values+="'"+i+"'"+', '
        final_values=final_values[:-2]
        print("Hello Boy",final_values)
        #conn.execute("PRAGMA foreign_keys = on")
        #query="INSERT INTO "+table+" VALUES("+final_values+");"
        #conn.execute(query)
        #conn.commit() 
        try:
            conn.execute("PRAGMA foreign_keys = on")
            query="INSERT INTO "+table+" VALUES("+final_values+");"
            conn.execute(query)
            conn.commit()
            conn.close()
            return Response(status=201)
        except:
            print("Where Am I now ??")    
            return Response(status=500)
     
    elif(OP == 2):
        column = request.get_json()['column']
        query = "UPDATE Count SET "+column+"="+column+"+1"+";"
        
        conn.execute(query)
        conn.commit()
        conn.close()
        return jsonify({}),200
    
    
    elif(OP==3):
        value = request.get_json()['value']
        conn.execute("UPDATE DLReg SET Flag=1,status=1 WHERE DLno=?;",(value,))
        conn.commit()
        conn.close()
        return Response(status=200)

    elif(OP==4):
        value = request.get_json()['value']
        conn.execute("UPDATE RCReg SET Approved=1 WHERE vahicle_no=?;",(value,))
        conn.commit()
        conn.close()
        return Response(status=200)
    elif(OP==5):
        value = request.get_json()['value']
        conn.execute("UPDATE LLReg SET Approved=1 WHERE reg_id=?;",(value,))
        conn.commit()
        conn.close()
        return Response(status=200)
    
   

@app.route("/api/v1/db/read",methods=["POST"])
@cross_origin(origin="*")
def readDB():
    OP = request.get_json()["OP"]
    try:
        conn = sqlite3.connect("webdatabase.db")
    except:
        return Response(status=500)

    if(OP == 1):# RETRIEVE DATA FROM ANY TABLE WITH COLUMN INFO 
        table = request.get_json()['table']
        column = request.get_json()['column']
        src = request.get_json()['src']
        dst = request.get_json()['dst']
        where = request.get_json()['where']
        
        try:
            if(where==1):
                res = conn.execute("SELECT "+column+" FROM "+table+" WHERE "+src+"=?",(dst,))
                final = res.fetchall()
                print("I am here :)",final)
                if(len(final)==0):
                    return Response(status=204)
            
                return json.dumps(final),200
            else:
                query="SELECT "+column+" FROM "+table+";"
                res = conn.execute(query)
                final = res.fetchall()
                print("Lo I am here ",final)	
                if(len(final)==0):
                    return Response(status=204)
            
                return json.dumps(final),200
        except:
            return Response(status=500)
    elif(OP == 2):
        table = request.get_json()['table']
        column = request.get_json()['column']
        src = request.get_json()['src']
        dst = request.get_json()['dst']
        where = request.get_json()['where']
         
        
        try:
            if(where==1):
               
                res = conn.execute("SELECT * FROM LLReg WHERE Name=? AND Mail=? AND DOB=?",(dst[0],dst[1],dst[2]))
                final = res.fetchall()
                print("I am here :)",final)
                if(len(final)==0):
                    return Response(status=204)
            
                return json.dumps(final),200
            else:
                query="SELECT "+column+" FROM "+table+";"
                res = conn.execute(query)
                final = res.fetchall()
                print("Lo I am here ",final)	
                if(len(final)==0):
                    return Response(status=204)
            
                return json.dumps(final),200
        except:
            return Response(status=500)
    elif(OP == 3):
        
        try:
            query = "SELECT RTO,SUM(SRevenue),SUM(VRevenue) FROM Revenue GROUP BY RTO;"
            res = conn.execute(query)
            final = res.fetchall()
            print(final)
            if(len(final)==0):
                return jsonify({}),204
            return json.dumps(final),200 
        except:
            return jsonify({}),500
    elif(OP == 4):
        
        try:
            query = "SELECT RTO,SUM(LLapplications),SUM(DLapplications),SUM(RCapplications) FROM Total GROUP BY RTO;"
            res = conn.execute(query)
            final = res.fetchall()
            print(final)
            if(len(final)==0):
                return jsonify({}),204
            return json.dumps(final),200 
        except:
            return jsonify({}),500
   


# coding: utf-8

# In[ ]:




# In[ ]:


from plotly.offline import plot
from plotly.graph_objs import Scatter
from flask import Markup


# In[ ]:


@app.route('/rcreg/<year>', methods=['GET','POST','OPTIONS'])
@cross_origin(origin="*")
def test(year):
    if request.method == 'GET':
        year=int(year)
        df=pd.read_excel('regdata.xlsx')
        fil=(df['year']== year)
        #df=df[df.year.eq(year)]
        df=df[fil]
        print(df)

        list1=df['month'].tolist()
        list2=df['count'].tolist()
        #my_plot_div = plot([Bar(x=list1, y=list2)], output_type='div')
        pio.renderers.default = "browser"
        fig = go.Figure([go.Bar(x=list1,y=list2)])
        fig.show()
        return "Graph Successful"
        #return render_template('test1.html',div_placeholder=Markup(my_plot_div))
@app.route('/license/appln', methods=['GET', 'POST'])
def applncomp():
    if request.method == 'GET':
        df=pd.read_excel('applcount.xlsx')
        list1=df['type'].tolist()
        list2=df['countofappl'].tolist()
        pio.renderers.default = "browser"
        fig = go.Figure([go.Pie(labels=list1,values=list2)])
        fig.show()
        return "Graph Successful"
@app.route('/vahanrev/<year>', methods=['GET', 'POST'])
def revenue(year):
    if request.method == 'GET':
        year=int(year)
        df=pd.read_excel('vahanrevenue.xlsx')
        fil=(df['year']== year)
        df=df[fil]
        print(df)
        list1=df['month'].tolist()
        list2=df['count'].tolist()
        pio.renderers.default = "browser"
        fig = go.Figure([go.Bar(x=list1,y=list2)])
        fig.show()
        
@app.route('/rcreg/cmpyear', methods=['GET', 'POST'])
def rcregcmpyear():
    if request.method == 'GET':
        df=pd.read_excel('regdata.xlsx')
        years=[2016,2017,2018,2019,2020]
        total=[]
        for year in years:
            fil=(df['year']== year)
            df_temp=df[fil]
            sumtotal=df_temp['count'].sum()
            total.append(sumtotal)
            
        print(total)
        pio.renderers.default = "browser"
        fig = go.Figure([go.Bar(x=years,y=total)])
        fig.show()
        
@app.route('/vahanrev/cmpyear', methods=['GET', 'POST'])
def revcmpyear():
    if request.method == 'GET':
        df=pd.read_excel('vahanrevenue.xlsx')
        years=[2016,2017,2018,2019,2020]
        total=[]
        for year in years:
            fil=(df['year']== year)
            df_temp=df[fil]
            sumtotal=df_temp['count'].sum()
            total.append(sumtotal)
            
        print(total)
        pio.renderers.default = "browser"
        fig = go.Figure([go.Bar(x=years,y=total)])
        fig.show()
        


    


# In[ ]:
# In[ ]:
'''

import pandas as pd
year=2016
df=pd.read_excel('regdata.xlsx')
fil=(df['year']== year)
df[fil]
list1=df['month'].tolist()
list2=df['count'].tolist()


# In[ ]:


list1
'''



@app.route('/api/search/<term>',methods=["GET"])
@cross_origin(origin="*")
def searchbar(term):
    file = open('search.txt', 'r') 
    content = file.read()
    lines = content.splitlines()
    listItem = []
    if(term != ''):
        for line in lines:
            if(term.upper() in line.upper()):
                listItem.append(line)
    response = app.response_class(response=json.dumps(listItem),status=200,mimetype='application/json')
    return response
    
@app.route('/api/getXML',methods=["GET"])
@cross_origin(origin="*")
def getXML():
    file = open('rss.xml', 'r') 
    content = file.read()
    response = app.response_class(response=json.dumps(content),status=200,mimetype='application/json')
    return response

@app.route("/api/v1/checkDLStatus",methods=["POST"])
@cross_origin(origin='*')
def checkdlStatus():
    if(request.method == 'POST'):
        DLno = request.get_json()["dlnum"]
        DOB = request.get_json()["DoB"]
        readresponse_dl = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"DLReg","column":"LLno,DOA,Flag,Status,Validity","src":"DLno","dst":DLno,"where":1})
        if(readresponse_dl.status_code == 204):
            return jsonify({}),400
        LLno = readresponse_dl.json()[0][0] 
        print(LLno)
        readresponse_ll = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"LLReg","column":"RTO,DOB,Class,Name","src":"reg_id","dst":LLno,"where":1})
        #print(readresponse_ll.status_code)
        if(readresponse_ll.status_code == 204):
            return jsonify({}),400
        if(readresponse_ll.json()[0][1] != DOB):
            return jsonify({}),400
        status = readresponse_dl.json()[0][3]
        name = readresponse_ll.json()[0][3]
        doi = readresponse_dl.json()[0][1]
        RTO = readresponse_ll.json()[0][0]
        validity = readresponse_dl.json()[0][4]
        cov = readresponse_ll.json()[0][2]
        if(status == 1):
            status = 'ACTIVE'
        else:
            status = 'NOT ACTIVE'
        found = 1
        if(found == 1): 
            result = {"found":1,"dlnum":DLno,"status":status,"name":name,"doi":doi,"RTO":RTO,"validity":validity,"cov":cov,"new":1}
        else:
            result = {"found":0}
        response = app.response_class(response=json.dumps(result),status=200,mimetype='application/json')
        return response
        
    else:
        return Response(status=405)

@app.route("/api/v1/checkRCStatus",methods=["POST"])
@cross_origin(origin='*')
def checkrcStatus():
    if(request.method == 'POST'):
        vehicle_no = request.get_json()["rcnum"]
        print(vehicle_no)
        readresponse = requests.post("http://127.0.0.1:5000/api/v1/db/read",json={"OP":1,"table":"RCReg","column":"RTO,DOR,Chassis,Engine,Owner,Class,Manufacturer,Approved","src":"vahicle_no","dst":vehicle_no,"where":1})
        if(readresponse.status_code == 204):
            return jsonify({}),400#Vehicle No does not exist.
        #print(readresponse.json())       
        
        
        RCno = request.get_json()["rcnum"]
        RTO = readresponse.json()[0][0]
        status = readresponse.json()[0][7]
        dor = readresponse.json()[0][1]
        chassis = readresponse.json()[0][2]
        engine = readresponse.json()[0][3]
        owner = readresponse.json()[0][4]
        cov = readresponse.json()[0][5]
        model = readresponse.json()[0][6]
        validity = '30-01-2028'
        if(status == 1):
            status = 'ACTIVE'
        else:
            status = 'INACTIVE'
        found = 1
        if(found == 1): 
            result = {"found":1,"rcnum":RCno,"status":status,"owner":owner,"dor":dor,"RTO":RTO,"validity":validity,"cov":cov,"chassis":chassis,"engine":engine,"model":model}
        else:
            result = {"found":0}
        response = app.response_class(response=json.dumps(result),status=200,mimetype='application/json')
        return response
        
    else:
        return Response(status=405)















if __name__ == '__main__': 
    app.secret_key = 'A0Zr98j/3yX R~XHH!jmN]LWX/,?RT'
    app.run(host='0.0.0.0',debug=True)

